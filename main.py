import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD
import pandas as pd
import os
import string
import re
from openpyxl import load_workbook
from openpyxl.styles import Font


class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor")
        self.label = tk.Label(root, text="Drag and Drop an Excel file here")
        self.label.pack(padx=10, pady=10)

        # Bind the drop event
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>', self.drop)

    def drop(self, event):
        file_path = event.data.strip('{}')
        self.process_file(file_path)
    def replace_special_chars(self, value):
        if pd.isna(value):
            return "Aucun"
        if isinstance(value, str) and len(value) == 1 and value in string.punctuation:
            return "Aucun"
        return value
    def replace_politique_voyage(self, value):
        if pd.isna(value):
            return "GENERAL.DEFAULT"
        if isinstance(value, str) and len(value) == 1 and value in string.punctuation:
            return "GENERAL.DEFAULT"
        return value
    def replace_genre(self, value):
        if isinstance(value, str) and re.search(r'ma', value, re.IGNORECASE):
            return "Mrs"
        elif isinstance(value, str) and re.search(r'mo', value, re.IGNORECASE):
            return "Mr"
        return value
    def first_name(self, value):
        if isinstance(value, str):
            return value.capitalize()
        return value
    def last_name(self, value):
        if isinstance(value, str):
            return value.upper()
        return value
    def role(self, value):
        if isinstance(value, str) and re.search(r'admin|executive', value, re.IGNORECASE):
            return "executive"
        elif isinstance(value, str) and re.search(r'mana|booker', value, re.IGNORECASE):
            return "booker"
        elif isinstance(value, str) and re.search(r'compta|accountant', value, re.IGNORECASE):
            return "accountant"
        elif isinstance(value, str) and re.search(r'voyage|traveler', value, re.IGNORECASE):
            return "traveler"
        if pd.isna(value):
            return "Vide"
        return value
    def langue(self, value):
        if isinstance(value, str) and re.search(r'fr', value, re.IGNORECASE):
            return "fr"
        elif isinstance(value, str) and re.search(r'an', value, re.IGNORECASE):
            return "en"
        elif isinstance(value, str) and re.search(r'en', value, re.IGNORECASE):
            return "en"
        elif isinstance(value, str) and re.search(r'es', value, re.IGNORECASE):
            return "es"
        elif isinstance(value, str) and re.search(r'sp', value, re.IGNORECASE):
            return "es"
        return value
    def date_de_naissance(self, value):
        try:
            return pd.to_datetime(value).strftime('%Y-%m-%d')
        except Exception:
            return value
    def clean_email(self, value):
        if pd.isna(value):
            return "Vide"
        if isinstance(value, str) and re.search(r'[ ,;/\']', value):
            return value
        return value
    # def clean_tel(self, value):
    #     if value == "nan":
    #         value = ""
    #     elif isinstance(value, str) and value.strip():  # Vérifie si la valeur est une chaîne non vide
    #         # Nettoyer le numéro de téléphone en supprimant les caractères spéciaux
    #         cleaned_value = re.sub(r'[ .+]', "", value)
    #         # Ajouter un '0' devant le numéro si ce n'est pas déjà présent
    #         if not cleaned_value.startswith('0'):
    #             cleaned_value = '0' + cleaned_value
    #         return cleaned_value
    #     return value

    def clean_tel(self, value):
        if value == "nan":
            value = ""
            return value
        elif isinstance(value, str) and value.strip():  # Vérifie si la valeur est une chaîne non vide
            # Nettoyer le numéro de téléphone en supprimant les caractères spéciaux
            cleaned_value = re.sub(r'[ .+]', "", value)
            # cleaned_value = "'" + cleaned_value
            # Ajouter un '0' devant le numéro si ce n'est pas déjà présent
            # if not cleaned_value.startswith('0'):
            #     cleaned_value = '0' + cleaned_value
            return cleaned_value

    def acces(self, value):
        if pd.isna(value):
            return ""
        if isinstance(value, str):
            value_lower = value.lower()
            if any(kw in value_lower for kw in ["oui", "v", "vrai", "true"]):
                return "true"
            elif any(kw in value_lower for kw in ["non", "f", "faux", "false"]):
                return "false"
        return value
    def clean_emails(self, value):
        if isinstance(value, str):
            # Remplace les caractères spéciaux par un espace unique
            return re.sub(r'[^a-zA-Z0-9@._]+', ' ', value).strip()
        return value
    def clear_column_if_not_empty(self, value):
        if isinstance(value, str) and value.strip():
            return ""
        return value
    def contains_special_chars_or_spaces(self, value):
        if isinstance(value, str) and re.search(r'[ ,;/\'\s]', value):
            return True
        return False
    def process_file(self, file_path):
        try:
            # Load the Excel file
            df = pd.read_excel(file_path)

            # Save the original dataframe for comparison
            original_df = df.copy()

            # Add the "TEST" column with "OK" values
            df['Centre de coût principal'] = df['Centre de coût principal'].str.upper()

            column_name = "Centre de coût secondaire / service"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.replace_special_chars)

            column_name = "Politique de voyage"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.replace_politique_voyage)

            column_name = "Genre"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.replace_genre)

            column_name = "Prénom"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.first_name)

            column_name = "Nom de famille"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.last_name)

            column_name = "Rôle"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.role)

            column_name = "Langue"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.langue)

            column_name = "Date de naissance"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.date_de_naissance)

            column_name = "Email"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.clean_email)

            df["TEL"] = df["TEL"].astype(str)
            if "TEL" in df.columns:
                df["TEL"] = df["TEL"].apply(self.clean_tel)

            columns_to_process = [
                "Sans accès",
                "Peut réserver pour lui sans validation dans la politique",
                "Peut réserver pour les autres sans validation",
                "Peut réserver pour lui sans validation hors politique",
                "Peut valider dans la politique",
                "Peut valider hors politique",
                "Peut voir les offres hors politique",
                "Validation RSE",
                "Recevoir les demandes de réservations des membres de l'équipe",
                "Recevoir les confirmations de réservations des membres de l'équipe",
                "Recevoir les reçus",
                "Recevoir les factures périodiques"
            ]
            for col in columns_to_process:
                if col in df.columns:
                    df[col] = df[col].apply(self.acces)

            column_name = "Assigner valideur (Manager ou Administrateur)"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.clean_emails)

            column_name = "Recevoir tout (admin)"
            if column_name in df.columns:
                df[column_name] = df[column_name].apply(self.clear_column_if_not_empty)

            # Ensure all specified columns are formatted as text
            for col in columns_to_process:
                if col in df.columns:
                    df[col] = df[col].astype(str)
                    df[col] = df[col].str.lower()

            # Construct the new file path
            dir_name = os.path.dirname(file_path)
            base_name = os.path.basename(file_path)
            new_file_path = os.path.join(dir_name, f"clean_{base_name}")

            # Save the modified DataFrame to a new Excel file
            df.to_excel(new_file_path, index=False)

            # Load the new file with openpyxl to apply styles
            wb = load_workbook(new_file_path)
            ws = wb.active

            font_green = Font(color="00B050")  # Define the font color green
            font_red = Font(color="FF0000")    # Define the font color red
            # Iterate over the DataFrame to check for changes and apply styles
            for row in range(2, df.shape[0] + 2):  # Adjust for header row in Excel
                for col in range(1, df.shape[1] + 1):
                    cell = ws.cell(row=row, column=col)
                    original_value = original_df.iloc[row-2, col-1]  # Adjust for header row in DataFrame
                    new_value = df.iloc[row-2, col-1]
                    if original_value != new_value:
                        cell.font = font_green  # Apply the green font to modified cells
                    if col in [df.columns.get_loc("Email")+1, df.columns.get_loc("Rôle")+1] and new_value == 'Vide':
                        cell.font = font_red  # Apply the red font to cells with "Vide"
                    if col in [df.columns.get_loc("Email") + 1] and self.contains_special_chars_or_spaces(new_value):
                        cell.font = font_red  # Apply the red font to cells with special characters or spaces in "Email"

            # Remove the second row (index 2) from the Excel sheet
            ws.delete_rows(2)

            # Save the styled workbook
            wb.save(new_file_path)

            self.label.config(text=f"File processed and saved as {new_file_path}")

        except Exception as e:
            self.label.config(text=f"Error: {e}")

if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()